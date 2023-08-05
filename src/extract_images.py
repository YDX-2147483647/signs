from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import TYPE_CHECKING
from zipfile import ZipFile

import cv2
import numpy as np
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image, ImageFilter
from util import homomorphic_filter

if TYPE_CHECKING:
    from typing import Generator, Literal


@dataclass
class Record:
    name: str
    image: Image.Image
    src: Literal["sign"] | Literal["photo"]


def load_records(
    table_filepath: Path, photos_filepath: Path
) -> Generator[Record, None, None]:
    workbook = load_workbook(table_filepath)
    sheet = workbook[workbook.sheetnames[0]]
    image_loader = SheetImageLoader(sheet)

    with ZipFile(photos_filepath) as photos_zip:
        for row in sheet.iter_rows(min_row=2, min_col=2):
            name: str = row[0].value  # type: ignore

            if (photo_filename := row[1].value) is not None:  # type: ignore
                photo_filename: str
                with photos_zip.open(photo_filename) as file:
                    photo = Image.open(BytesIO(file.read()))
            else:
                photo = None

            if image_loader.image_in(row[2].coordinate):
                sign = image_loader.get(row[2].coordinate)
            else:
                sign = None

            assert (
                sum(i is not None for i in (photo, sign)) == 1
            ), f"“{name}”没有正常提供照片或签名：{photo = }，{sign = }"

            yield Record(
                name=name,
                image=photo or sign,  # type: ignore
                src="photo" if photo else "sign",
            )


def is_transparent(image: Image.Image) -> bool:
    if "A" not in image.getbands():
        # e.g. RGB
        return False
    else:
        # e.g. RGBA
        alpha = image.getchannel("A")
        return (np.asarray(alpha) < 255).any()  # type: ignore


def transparentize_background(image: Image.Image) -> None:
    threshold, img = cv2.threshold(
        np.asarray(image.convert("L")), 0, 255, cv2.THRESH_OTSU
    )
    image.putalpha(
        image.convert("L").point(lambda x: 255 if x < 1.2 * threshold else 0)
    )


def auto_crop(image: Image.Image) -> Image.Image:
    alpha = image.split()[-1]
    # 扩散会抑制噪声，添加边距
    nonzero = (np.asarray(alpha.filter(ImageFilter.BoxBlur(50))) > 10).nonzero()
    (upper, lower), (left, right) = [(min(x), max(x)) for x in nonzero]
    return image.crop((left, upper, right + 1, lower + 1))


if __name__ == "__main__":
    table_filepath = next((Path.home() / "Downloads").glob("*.xlsx"))
    photos_filepath = next((Path.home() / "Downloads").glob("*.zip"))
    out_dir = Path("out")

    out_dir.mkdir(exist_ok=True)

    records = load_records(table_filepath, photos_filepath)
    for r in records:
        # 删除背景
        should_remove_background = r.src == "photo" and not is_transparent(r.image)
        if should_remove_background:
            print(f"正在删除“{r.name}”的背景。")
            transparentize_background(r.image)

        # 裁切
        print(f"正在裁切“{r.name}”。")
        r.image = auto_crop(r.image)

        # 检查
        if should_remove_background:
            gray = np.asarray(r.image.convert("L"))
            if gray.std() < 30:
                print(f"“{r.name}”的对比度太低，正在同态滤波，准备重试。")
                gray = homomorphic_filter(gray, cutoff=60)
                print(f"正在重新删除“{r.name}”的背景。")
                threshold, img = cv2.threshold(
                    gray, 0, 255, cv2.THRESH_OTSU | cv2.THRESH_BINARY_INV
                )
                r.image.putalpha(Image.fromarray(img))

        r.image.save(out_dir / f"{r.name}.png")
